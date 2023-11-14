#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Mon Nov 13 22:01:15 2023

@author: marik
"""

import pandas as pd
import seaborn as sns
import numpy as np
import matplotlib.pyplot as plt

# import xlrd
# file="patrik.xls"
# wb = xlrd.open_workbook(file, formatting_info=True)
# sheets = wb.sheet_names()

# import openpyxl
from openpyxl import load_workbook
file = 'Vysledky_Frekvence_22_03_2021_FSoriginal.xlsx' 
wb = load_workbook(file, data_only = True)
sheets = wb.sheetnames

# %%

df_optika = pd.read_excel("fft_data_01_Mereni_Babice_22032021_optika_zpracovani.xlsx", index_col=0)
df_optika.index = df_optika.index.str.split('_', n=1, expand=True)

for sheet in sheets[:-1]:
    print(file, sheet)
    
    df = pd.read_excel(file, sheet_name=sheet, index_col=0, header=[0,1,2])
    df.dropna(axis = 0, how = 'all', inplace = True)
    
    sheet_ = wb[sheet]
    bgcol = np.empty([sheet_.max_row,sheet_.max_column], dtype=object)
    for row in range(sheet_.max_row):
        print()
        for col in range(sheet_.max_column):
            c=sheet_.cell(row+1, col+1)
            bgcol[row,col] = c.fill.start_color.index
    pocet_mereni = df.shape[0]        
    bgcol = bgcol[3:3+pocet_mereni,1:]        
  
    df = pd.melt(df,  ignore_index=False)
    df.reset_index(inplace=True)
    df["barva"] = bgcol.T.reshape(-1)
    df.columns = ["mereni","pristroj","osa","metoda","frekvence","spolehlivost"]
    df['spolehlivost'] = df['spolehlivost'].astype('category').cat.rename_categories({"FFFFC000":"pochybny", "FFFF0000": "spatne", "00000000":"OK"})
    df['pristroj'] = df['pristroj'].str.replace(r' - .*', '', regex=True).astype('category')
        
    fig, axs = plt.subplots(2,2, figsize=(14,10))
    axs = axs.reshape(-1)
    plt.suptitle(f"{sheet}")
    for n,i in enumerate(df_optika.loc[sheet,"Freq"]):
        axs[0].axhline(i, label='optika')
        axs[2].plot([n],[i],"d", color='k', label='optika')
    
    for ax,x in zip(axs,["metoda","osa","mereni"]):
        sns.swarmplot(data=df, x=x, y="frekvence", ax=ax, hue="pristroj", dodge=True)
        ax.set(ylim=(-0.02,1))
        ax.grid(which="both")
    
    # neopakovat polozky v legende
    for ax in axs[0:3]:
        handles, labels = ax.get_legend_handles_labels()
        by_label = dict(zip(labels, handles))
        ax.legend(by_label.values(), by_label.keys())        
    
        
    ax = axs[3]    
    sns.swarmplot(data=df, y="frekvence", x="osa", hue="spolehlivost", ax=ax, dodge=True)    
    ax.set(ylim=(-0.02,1))
    ax.grid(which="both")
    plt.tight_layout(pad=2)
    plt.savefig(f"optika_plus_acc/{sheet}.png")

# %%


